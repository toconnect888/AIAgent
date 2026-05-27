"""Upload contacts from Excel sheet 3 (Flagged for Review) to HubSpot CRM."""

import time
import requests
import openpyxl
from datetime import datetime
from dotenv import load_dotenv
import os

load_dotenv()
api_key = os.environ.get("HUBSPOT_API_KEY")

EXCEL_PATH = "output/leads_2026-05-26.xlsx"
SHEET_NAME = "Flagged for Review"
STATUS_COL_HEADER = "HubSpot Upload Status"

HEADERS = {
    "Authorization": f"Bearer {api_key}",
    "Content-Type": "application/json",
}


def search_contact_by_email(email):
    url = "https://api.hubapi.com/crm/v3/objects/contacts/search"
    payload = {
        "filterGroups": [{"filters": [{"propertyName": "email", "operator": "EQ", "value": email}]}],
        "properties": ["email", "firstname", "lastname", "company", "phone"],
    }
    r = requests.post(url, headers=HEADERS, json=payload)
    r.raise_for_status()
    results = r.json().get("results", [])
    return results[0]["id"] if results else None


def create_contact(props):
    url = "https://api.hubapi.com/crm/v3/objects/contacts"
    r = requests.post(url, headers=HEADERS, json={"properties": props})
    r.raise_for_status()
    return r.json()["id"]


def update_contact(contact_id, props):
    url = f"https://api.hubapi.com/crm/v3/objects/contacts/{contact_id}"
    r = requests.patch(url, headers=HEADERS, json={"properties": props})
    r.raise_for_status()
    return contact_id


def add_note(contact_id, note_body):
    # Create note
    url = "https://api.hubapi.com/crm/v3/objects/notes"
    ts = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    r = requests.post(url, headers=HEADERS, json={"properties": {"hs_note_body": note_body, "hs_timestamp": ts}})
    r.raise_for_status()
    note_id = r.json()["id"]

    # Associate note with contact
    assoc_url = f"https://api.hubapi.com/crm/v3/objects/notes/{note_id}/associations/contacts/{contact_id}/note_to_contact"
    requests.put(assoc_url, headers=HEADERS)


def parse_name(full_name):
    parts = (full_name or "").strip().split(" ", 1)
    return parts[0], parts[1] if len(parts) > 1 else ""


def upload_sheet(ws):
    headers = [cell.value for cell in ws[1]]

    # Find or add status column
    if STATUS_COL_HEADER not in headers:
        status_col_idx = len(headers) + 1
        ws.cell(row=1, column=status_col_idx, value=STATUS_COL_HEADER)
    else:
        status_col_idx = headers.index(STATUS_COL_HEADER) + 1

    col = {h: i + 1 for i, h in enumerate(headers)}

    uploaded = 0
    skipped = 0

    for row_idx in range(2, ws.max_row + 1):
        def v(name):
            idx = col.get(name)
            return ws.cell(row=row_idx, column=idx).value if idx else None

        name = v("Name") or ""
        email = v("Email")
        company = v("Company")
        phone = str(v("Phone") or "")
        intent = v("Intent Summary") or ""
        reason = v("Reason Flagged") or ""

        if not email:
            ws.cell(row=row_idx, column=status_col_idx, value="SKIPPED – no email")
            skipped += 1
            continue

        firstname, lastname = parse_name(name)

        props = {
            "firstname": firstname,
            "lastname": lastname,
            "email": email,
            "phone": phone,
            "hs_lead_status": "NEW",
            "lifecyclestage": "lead",
        }
        # Only set company if it looks real
        if company and "unknown" not in company.lower() and "gmail" not in company.lower():
            props["company"] = company

        note_body = f"Intent: {intent}"
        if reason:
            note_body += f"\n\nFlagged reason: {reason}"

        try:
            existing_id = search_contact_by_email(email)
            if existing_id:
                update_contact(existing_id, props)
                add_note(existing_id, note_body)
                status = f"UPDATED – id {existing_id}"
                print(f"  Updated existing contact {email} (id {existing_id})")
            else:
                new_id = create_contact(props)
                add_note(new_id, note_body)
                status = f"CREATED – id {new_id}"
                print(f"  Created new contact {email} (id {new_id})")

            ws.cell(row=row_idx, column=status_col_idx, value=status)
            uploaded += 1
        except requests.HTTPError as e:
            err = f"ERROR – {e.response.status_code}: {e.response.text[:120]}"
            ws.cell(row=row_idx, column=status_col_idx, value=err)
            print(f"  FAILED for {email}: {err}")

        time.sleep(0.2)  # stay well under rate limit

    return uploaded, skipped


def main():
    print(f"Loading {EXCEL_PATH} …")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        return

    ws = wb[SHEET_NAME]
    print(f"Processing sheet: '{SHEET_NAME}' ({ws.max_row - 1} contact rows)")

    uploaded, skipped = upload_sheet(ws)

    wb.save(EXCEL_PATH)
    print(f"\nDone. Uploaded: {uploaded} | Skipped: {skipped}")
    print(f"Status written back to {EXCEL_PATH}")


if __name__ == "__main__":
    main()
